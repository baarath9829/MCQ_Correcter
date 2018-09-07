import openpyxl
from Tkinter import *
import string

def onsubmit(event):
    mark_register ={}
    filename = txtbox_filename.get()
    filevalidateflag = validatefilename(filename)
    if (filevalidateflag):
        wb = openpyxl.load_workbook(filename)
        master_key = []
        stringbuffer = ""
        mark = 0
        noofquestions = txtbox_noofquestions.get()
        numbervalidateflag = validatenumber(noofquestions)
        if (not(numbervalidateflag)):
            noofquestions = 0
        else:
            noofquestions = int(noofquestions)
        sheets =  wb.sheetnames
        if ("answerkey" in sheets):
            keysheet  = wb["answerkey"]
            for i in range(1,noofquestions+1):
                stringbuffer = stringbuffer + str(keysheet["A" + str(i)].value)
                stringbuffer = stringbuffer + str(keysheet["B" + str(i)].value)
                stringbuffer = stringbuffer + str(keysheet["C" + str(i)].value)
                stringbuffer = stringbuffer + str(keysheet["D" + str(i)].value)
                master_key.append(stringbuffer)
                stringbuffer = ""
        

        for sheet in sheets:
            print (sheet)
            if ((sheet != "answerkey") and (sheet != "result")):
                answersheet = wb[sheet]
                for i in range(1,noofquestions+1):
                    stringbuffer = stringbuffer + str(answersheet["A" + str(i)].value)
                    stringbuffer = stringbuffer + str(answersheet["B" + str(i)].value)
                    stringbuffer = stringbuffer + str(answersheet["C" + str(i)].value)
                    stringbuffer = stringbuffer + str(answersheet["D" + str(i)].value)
                    if (stringbuffer == master_key[i-1]):
                        mark = mark + 1
                    stringbuffer = ""
                answersheet["F1"].value = "Total :"
                answersheet["G1"].value = mark
                print (str(sheet)+ ":" +str(mark))
                mark_register[sheet]=mark
                wb.save(filename)
                mark = 0

        if ("result" not in sheets):
            wb.create_sheet("result")
            resultsheet = wb["result"]
            index = 1
            for sheet in sheets:
                if ((sheet != "answerkey") and (sheet != "result")):
                    resultsheet["A" + str(index)].value = sheet
                    resultsheet["B" + str(index)].value = mark_register[sheet]
                    index = index + 1
        else:
            print ("result sheet is already available")
        wb.save(filename)
def validatefilename(filename):
    if(filename != ""):
        if (type(filename)== str):
            if (".xlsx" in filename):
                return True
            else:
                print ("extension is invalid")
                return False
        else:
            print ("filename should be string")
            return False
    else:
        print ("text box is empty")
        return False
        
       
def validatenumber(number):
    checklist = list(string.ascii_lowercase)
    checklist+list(string.ascii_uppercase)
    checklist+list(string.punctuation)
    if(number != ""):
        flag = True
        for letter in checklist:
            if(letter in number):
                flag = False
        if (flag):
            flag = False
            for letter in range(10):
                if(str(letter) in number):
                    flag = True
            if (flag):
                return True
            else:
                print ("there is no number")
                return False
        else:
            print ("there should not any special character or alphabet")
            return False
    else:
        print("text box is empty")
        return False
    
window =  Tk()
Label(window , text="file name with extension").grid(row=0 , sticky=W)
txtbox_filename = Entry(window, width=20)
txtbox_filename.grid(row=0,column=1, sticky=E)
Label(window , text="number of question").grid(row=1 , sticky=W)
txtbox_noofquestions = Entry(window, width=20)
txtbox_noofquestions.grid(row=1,column=1, sticky=E)
button = Button(window , text="submit")
button.grid(row=2,column=1, sticky=E)
button.bind("<Button-1>" , onsubmit)
window.mainloop()
        
                
